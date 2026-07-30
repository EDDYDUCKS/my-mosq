import openpyxl
from django.contrib.auth import authenticate, get_user_model
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework import permissions
from rest_framework import parsers
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action, api_view, permission_classes, renderer_classes
from rest_framework.renderers import BaseRenderer, JSONRenderer
import zoneinfo
from django.db import transaction
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.utils import timezone

class ExcelBinaryRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data
from .models import Estudiante, Equipo, Prestamo, Sancion, BitacoraAccion
from .serializers import EstudianteSerializer, EquipoSerializer, PrestamoSerializer, SancionSerializer, BitacoraAccionSerializer
from .utils import registrar_auditoria, enviar_notificacion_email
import os
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
from django.conf import settings

User = get_user_model()


# ── Lista blanca de cuentas administradoras ────────────────────────────────
def _get_admin_emails() -> set[str]:
    """
    Lee ADMIN_EMAILS del entorno y retorna un set de correos en minúsculas.
    Siempre incluye al superusuario 'admin' como salvaguarda.
    """
    raw = os.getenv('ADMIN_EMAILS', '')
    emails = {e.strip().lower() for e in raw.split(',') if e.strip()}
    return emails


def _sync_admin_flag(user) -> None:
    """
    - Si el correo del usuario está en la lista blanca → is_staff = True
    - Si NO está → is_staff = False (a menos que sea superusuario)
    Solo guarda si hubo cambio.
    """
    if user.is_superuser:
        return                          # nunca tocar al superusuario
    allowed = _get_admin_emails()
    should_be_staff = user.email.lower() in allowed
    if user.is_staff != should_be_staff:
        user.is_staff = should_be_staff
        user.save(update_fields=['is_staff'])


class IsAdminOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return bool(request.user and request.user.is_staff)


class LoginAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        identifier = request.data.get('email') or request.data.get('username')
        password = request.data.get('password')

        if not identifier or not password:
            return Response({'detail': 'Email/usuario y contraseña son obligatorios.'}, status=400)

        username = identifier
        if '@' in identifier:
            user_by_email = User.objects.filter(email__iexact=identifier).first()
            if user_by_email:
                username = user_by_email.username

        user = authenticate(request, username=username, password=password)
        if not user:
            return Response({'detail': 'Credenciales inválidas.'}, status=401)

        # Sincronizar flag de admin con la lista blanca
        _sync_admin_flag(user)

        token, _ = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user': {
                'id': str(user.id),
                'email': user.email,
                'name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'role': 'admin' if (user.is_staff or user.is_superuser) else 'student',
            },
        })


class CurrentUserAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # Check if student needs to complete profile
        requiere_perfil = False
        if not user.is_staff and not user.is_superuser:
            if user.email and user.email.endswith('@est.ulsa.edu.ni'):
                if not user.carnet or not user.carrera or not user.ano_cursado:
                    requiere_perfil = True
        
        return Response({
            'id': str(user.id),
            'email': user.email,
            'name': f"{user.first_name} {user.last_name}".strip() or user.username,
            'role': 'admin' if (user.is_staff or user.is_superuser) else 'student',
            'requiere_completar_perfil': requiere_perfil,
        })

# ==========================================
# 1. VISTAS AUTOMÁTICAS DE LA API (CRUD)
# ==========================================

from rest_framework import filters

class EstudianteViewSet(viewsets.ModelViewSet):
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'username', 'email', 'carnet']

class EquipoViewSet(viewsets.ModelViewSet):
    queryset = Equipo.objects.all()
    serializer_class = EquipoSerializer
    permission_classes = [IsAdminOrReadOnly]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]

    def perform_create(self, serializer):
        equipo = serializer.save()
        registrar_auditoria(
            usuario=self.request.user,
            accion='CREAR_EQUIPO',
            descripcion=f"Equipo '{equipo.nombre}' creado (Total: {equipo.cantidad_total}, Disponibles: {equipo.cantidad_disponible}, Mantenimiento: {equipo.cantidad_mantenimiento})",
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        equipo = serializer.save()
        registrar_auditoria(
            usuario=self.request.user,
            accion='EDITAR_EQUIPO',
            descripcion=f"Equipo #{equipo.id} '{equipo.nombre}' actualizado (Total: {equipo.cantidad_total}, Disponibles: {equipo.cantidad_disponible}, Mantenimiento: {equipo.cantidad_mantenimiento})",
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        nombre = instance.nombre
        eq_id = instance.id
        instance.delete()
        registrar_auditoria(
            usuario=self.request.user,
            accion='ELIMINAR_EQUIPO',
            descripcion=f"Equipo #{eq_id} '{nombre}' eliminado del inventario",
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

class PrestamoViewSet(viewsets.ModelViewSet):
    queryset = Prestamo.objects.all()
    serializer_class = PrestamoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = (
            Prestamo.objects.all()
            .select_related('estudiante', 'entregado_por', 'recibido_por')
            .prefetch_related('detalles__equipo')
            .order_by('-fecha_prestamo', '-id')
        )
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(estudiante=self.request.user)

    def perform_create(self, serializer):
        if self.request.user.is_staff:
            estado = serializer.validated_data.get('estado')
            save_kwargs = {}
            if estado == 'ACTIVO':
                save_kwargs['entregado_por'] = self.request.user
            
            # Check sanction if assigning to a student
            estudiante = serializer.validated_data.get('estudiante')
            if estudiante and estudiante.sancionado:
                raise PermissionDenied(f'El estudiante {estudiante.first_name} tiene una sanción activa y no puede recibir préstamos.')
                
            serializer.save(**save_kwargs)
            return

        estudiante = serializer.validated_data.get('estudiante')
        if not estudiante or estudiante.id != self.request.user.id:
            raise PermissionDenied('Solo puedes crear préstamos para tu propio usuario.')

        if self.request.user.sancionado:
            raise PermissionDenied('No puedes solicitar préstamos porque tienes una sanción activa.')

        # Validar Horario de Atención de Bodega (Lunes a Sábado, 7:00 AM - 7:00 PM, UTC-6)
        try:
            nicaragua_tz = zoneinfo.ZoneInfo('America/Managua')
            now_ni = timezone.now().astimezone(nicaragua_tz)
            if now_ni.isoweekday() == 7 or now_ni.hour < 7 or now_ni.hour >= 19:
                raise PermissionDenied('La bodega de deportes está cerrada en este momento. Horario de atención: Lunes a Sábado de 7:00 AM a 7:00 PM.')
        except Exception as e:
            if isinstance(e, PermissionDenied):
                raise e

        # Solo los estudiantes (@est.ulsa.edu.ni) deben tener carnet y carrera
        # Los profesores (@ac.ulsa.edu.ni) y staff (@ulsa.edu.ni) pueden prestar sin esos datos
        es_estudiante = self.request.user.email.lower().endswith('@est.ulsa.edu.ni')
        if es_estudiante and (not self.request.user.carnet or not self.request.user.carrera):
            raise PermissionDenied('Debes completar tu perfil (carnet y carrera) antes de solicitar equipos.')

        serializer.save(estado='PENDIENTE')

    def perform_update(self, serializer):
        estado_actual = serializer.instance.estado
        nuevo_estado = self.request.data.get('estado', estado_actual)

        # Permitir a estudiantes cancelar sus propios préstamos pendientes
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            if (serializer.instance.estudiante == self.request.user
                    and estado_actual == 'PENDIENTE'
                    and nuevo_estado == 'RECHAZADO'):
                serializer.instance.estado = 'RECHAZADO'
                serializer.instance.save(update_fields=['estado'])
                return
            raise PermissionDenied('Solo administradores pueden actualizar préstamos.')

        save_kwargs = {}

        if nuevo_estado == 'ACTIVO' and estado_actual != 'ACTIVO':
            p_inst = serializer.instance
            # Verificación Anti-Overbooking
            for detalle in p_inst.detalles.select_related('equipo').all():
                eq = detalle.equipo
                disp = eq.recalcular_disponibilidad()
                if disp < detalle.cantidad:
                    raise ValidationError(f"No hay suficiente stock disponible de '{eq.nombre}' (Quedan {disp} disponibles, solicita {detalle.cantidad}).")

            save_kwargs['entregado_por'] = self.request.user
            registrar_auditoria(
                usuario=self.request.user,
                accion='APROBAR_PRESTAMO',
                descripcion=f"Préstamo #{p_inst.id} APROBADO/ENTREGADO a {p_inst.estudiante.username}",
                ip_address=self.request.META.get('REMOTE_ADDR')
            )
            enviar_notificacion_email(
                destinatario_email=p_inst.estudiante.email,
                asunto=f"Préstamo Aprobado #{p_inst.id} - ULSA",
                mensaje_texto=f"Hola {p_inst.estudiante.first_name}, tu solicitud de préstamo #{p_inst.id} ha sido APROBADA. Puedes retirar tu equipo en la bodega de deportes."
            )

        if nuevo_estado == 'DEVUELTO' and estado_actual != 'DEVUELTO':
            save_kwargs['recibido_por'] = self.request.user
            save_kwargs['fecha_recepcion'] = timezone.now()
            p_inst = serializer.instance
            registrar_auditoria(
                usuario=self.request.user,
                accion='RECIBIR_PRESTAMO',
                descripcion=f"Préstamo #{p_inst.id} RECIBIDO (Devuelto) de {p_inst.estudiante.username}",
                ip_address=self.request.META.get('REMOTE_ADDR')
            )
            enviar_notificacion_email(
                destinatario_email=p_inst.estudiante.email,
                asunto=f"Constancia de Devolución Préstamo #{p_inst.id} - ULSA",
                mensaje_texto=f"Hola {p_inst.estudiante.first_name}, tu préstamo #{p_inst.id} ha sido entregado en bodega y marcado como DEVUELTO exitosamente."
            )
        elif estado_actual == 'DEVUELTO' and nuevo_estado != 'DEVUELTO':
            save_kwargs['recibido_por'] = None
            save_kwargs['fecha_recepcion'] = None

        if nuevo_estado == 'RECHAZADO' and estado_actual != 'RECHAZADO':
            p_inst = serializer.instance
            registrar_auditoria(
                usuario=self.request.user,
                accion='RECHAZAR_PRESTAMO',
                descripcion=f"Préstamo #{p_inst.id} RECHAZADO para {p_inst.estudiante.username}",
                ip_address=self.request.META.get('REMOTE_ADDR')
            )
            enviar_notificacion_email(
                destinatario_email=p_inst.estudiante.email,
                asunto=f"Solicitud de Préstamo Rechazada #{p_inst.id} - ULSA",
                mensaje_texto=f"Hola {p_inst.estudiante.first_name}, tu solicitud de préstamo #{p_inst.id} ha sido rechazada."
            )

        serializer.save(**save_kwargs)
        # Recalcular disponibilidades
        for detalle in serializer.instance.detalles.select_related('equipo').all():
            detalle.equipo.recalcular_disponibilidad()

    @action(detail=True, methods=['post'])
    def declarar_perdido(self, request, pk=None):
        """Marca un préstamo como PERDIDO, descuenta patrimonio total y sanciona al estudiante."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied('Solo administradores pueden declarar equipos perdidos.')

        prestamo = self.get_object()
        if prestamo.estado == 'PERDIDO':
            return Response({'detail': 'Este préstamo ya fue marcado como perdido.'}, status=400)

        motivo = request.data.get('motivo', 'Equipo extraviado / no devuelto')

        with transaction.atomic():
            prestamo.estado = 'PERDIDO'
            prestamo.observaciones = f"[EQUIPO PERDIDO] {motivo}"
            prestamo.save(update_fields=['estado', 'observaciones'])

            for detalle in prestamo.detalles.select_related('equipo').all():
                eq = detalle.equipo
                eq.cantidad_total = max(0, eq.cantidad_total - detalle.cantidad)
                eq.save(update_fields=['cantidad_total'])
                eq.recalcular_disponibilidad()

            sancion = Sancion.objects.create(
                estudiante=prestamo.estudiante,
                creada_por=request.user,
                severity='restriction',
                reason=f"Reposición por equipo perdido en Solicitud #{prestamo.id}: {motivo}",
                notes="Sanción automática por extravío de patrimonio universitario."
            )

            registrar_auditoria(
                usuario=request.user,
                accion='EDITAR_EQUIPO',
                descripcion=f"Préstamo #{prestamo.id} declarado como PERDIDO para {prestamo.estudiante.username}. Se descontó stock total y se generó sanción.",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            enviar_notificacion_email(
                destinatario_email=prestamo.estudiante.email,
                asunto=f"Notificación de Equipo Extraviado #{prestamo.id} - ULSA",
                mensaje_texto=f"Hola {prestamo.estudiante.first_name}, tu préstamo #{prestamo.id} ha sido reportado como EXTRAVIADO/PERDIDO. Se ha registrado una sanción por reposición en tu cuenta."
            )

        return Response(PrestamoSerializer(prestamo).data)

    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Permite al estudiante cancelar su propio préstamo PENDIENTE."""
        prestamo = self.get_object()
        
        # Solo el dueño puede cancelar
        if prestamo.estudiante != request.user:
            raise PermissionDenied('Solo puedes cancelar tus propios préstamos.')
        
        # Solo se pueden cancelar préstamos pendientes
        if prestamo.estado != 'PENDIENTE':
            return Response(
                {'detail': 'Solo se pueden cancelar préstamos en estado PENDIENTE.'},
                status=400
            )
        
        prestamo.estado = 'RECHAZADO'
        prestamo.save(update_fields=['estado'])
        
        return Response({'detail': 'Préstamo cancelado exitosamente.'})

    @action(detail=False, methods=['post'])
    def procesar_atrasados(self, request):
        if not request.user.is_staff:
            raise PermissionDenied('Solo administradores pueden procesar préstamos atrasados.')
            
        hoy = timezone.localdate()
        
        # Buscar préstamos ACTIVOS cuya fecha_devolucion (solo fecha) sea menor a hoy
        # Como fecha_devolucion es DateTimeField, comparamos su fecha.
        from django.db import transaction
        
        prestamos_atrasados = Prestamo.objects.filter(
            estado='ACTIVO', 
            fecha_devolucion__date__lt=hoy
        )
        
        contador = 0
        with transaction.atomic():
            for prestamo in prestamos_atrasados:
                prestamo.estado = 'ATRASADO'
                prestamo.save(update_fields=['estado'])
                
                # Crear la sanción automática
                from .models import Sancion
                Sancion.objects.create(
                    estudiante=prestamo.estudiante,
                    motivo=f'Devolución tardía automática del Ticket #{prestamo.id}',
                    observaciones='El sistema ha detectado que la fecha límite de devolución ha expirado.',
                    severidad='restriction',
                    activa=True
                )
                contador += 1
                
        return Response({'detail': f'Se procesaron {contador} préstamos atrasados.'})


class SancionViewSet(viewsets.ModelViewSet):
    queryset = Sancion.objects.all().select_related('estudiante', 'creada_por', 'resuelta_por')
    serializer_class = SancionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Sancion.objects.all().select_related('estudiante', 'creada_por', 'resuelta_por')
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(estudiante=self.request.user)

    def _ensure_admin(self):
        if not self.request.user.is_staff:
            raise PermissionDenied('Solo administradores pueden gestionar sanciones.')

    def perform_create(self, serializer):
        self._ensure_admin()
        serializer.save(creada_por=self.request.user)

    def perform_update(self, serializer):
        self._ensure_admin()
        serializer.save()

    def perform_destroy(self, instance):
        self._ensure_admin()
        instance.delete()

    @action(detail=True, methods=['patch'])
    def resolver(self, request, pk=None):
        self._ensure_admin()
        sancion = self.get_object()
        sancion.activa = False
        sancion.resuelta_por = request.user
        sancion.fecha_resolucion = timezone.now()

        observaciones = request.data.get('observaciones')
        if observaciones is not None:
            sancion.observaciones = observaciones

        sancion.save()
        return Response(self.get_serializer(sancion).data)


class BitacoraViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = BitacoraAccion.objects.all()
    serializer_class = BitacoraAccionSerializer
    permission_classes = [IsAdminUser]


# ==========================================
# 2. GENERADOR DE REPORTES EXCEL (VERSIÓN CARRITO)
# ==========================================

@api_view(['GET'])
@permission_classes([IsAdminUser])
@renderer_classes([ExcelBinaryRenderer, JSONRenderer])
def exportar_reporte_excel(request):
    import calendar
    import logging
    from datetime import datetime as dt
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    logger = logging.getLogger(__name__)

    try:
        # ── Colores institucionales ULSA ──────────────────────────────────────
        COLOR_ULSA        = '27822e'   # verde ULSA (fondo titulo y encabezados)
        COLOR_HEADER_TEXT = 'FFFFFF'   # blanco para texto de encabezados
        COLOR_FILA_PAR    = 'F0F7F0'   # verde muy claro para filas pares
        COLOR_FILA_IMPAR  = 'FFFFFF'   # blanco para filas impares
        # Colores por estado de préstamo
        COLOR_DEVUELTO    = 'C8E6C9'   # verde claro
        COLOR_PENDIENTE   = 'FFF9C4'   # amarillo claro
        COLOR_ACTIVO      = 'FFF9C4'   # amarillo claro
        COLOR_ATRASADO    = 'FFCDD2'   # rojo claro
        COLOR_RECHAZADO   = 'EEEEEE'   # gris claro

        # ── Filtro por mes (parámetro ?mes=2026-07) ───────────────────────────
        mes_param = request.GET.get('mes', None)
        hoy = timezone.localtime(timezone.now())
        if mes_param:
            try:
                anio, mes_num = map(int, mes_param.split('-'))
            except (ValueError, AttributeError):
                anio, mes_num = hoy.year, hoy.month
        else:
            anio, mes_num = hoy.year, hoy.month

        nombre_mes = calendar.month_name[mes_num]
        MESES_ES = {
            'January':'Enero','February':'Febrero','March':'Marzo','April':'Abril',
            'May':'Mayo','June':'Junio','July':'Julio','August':'Agosto',
            'September':'Septiembre','October':'Octubre','November':'Noviembre','December':'Diciembre'
        }
        nombre_mes_es = MESES_ES.get(nombre_mes, nombre_mes)

        # ── Carrera: codigo → nombre completo ────────────────────────────────
        CARRERAS_MAP = {
            'LAF': 'Lic. Administrativa - Finanzas',
            'LCM': 'Lic. Comercial - Mercadeo',
            'IGI': 'Ing. Gestión Industrial',
            'ICE': 'Ing. Cibernética Electrónica',
            'IME': 'Ing. Mecánica y Energías Renovables',
            'IMS': 'Ing. Mecatrónica y Sistemas',
            'IEM': 'Ing. Electromédica',
        }

        # ── Todos los préstamos del mes seleccionado ─────────────────────────
        tz_local = timezone.get_current_timezone()
        primer_dia = timezone.make_aware(dt(anio, mes_num, 1, 0, 0, 0), tz_local)
        ultimo_dia_num = calendar.monthrange(anio, mes_num)[1]
        ultimo_dia = timezone.make_aware(dt(anio, mes_num, ultimo_dia_num, 23, 59, 59), tz_local)

        prestamos = (
            Prestamo.objects
            .filter(fecha_prestamo__range=(primer_dia, ultimo_dia))
            .select_related('estudiante', 'entregado_por', 'recibido_por')
            .prefetch_related('detalles__equipo')
            .order_by('-fecha_prestamo')
        )

        # Conteo de estados para resumen KPI
        cant_devueltos  = prestamos.filter(estado='DEVUELTO').count()
        cant_activos    = prestamos.filter(estado='ACTIVO').count()
        cant_atrasados  = prestamos.filter(estado='ATRASADO').count()
        cant_pendientes = prestamos.filter(estado='PENDIENTE').count()
        cant_rechazados = prestamos.filter(estado='RECHAZADO').count()

        # ── Crear workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        
        # Estilos reutilizables
        fill_ulsa    = PatternFill(fill_type='solid', start_color=COLOR_ULSA, end_color=COLOR_ULSA)
        fill_header  = PatternFill(fill_type='solid', start_color='1A6B20', end_color='1A6B20')
        font_white14 = Font(name='Calibri', bold=True, size=14, color=COLOR_HEADER_TEXT)
        font_white11 = Font(name='Calibri', bold=True, size=11, color=COLOR_HEADER_TEXT)
        font_bold    = Font(name='Calibri', bold=True, size=10)
        font_normal  = Font(name='Calibri', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        thin = Side(border_style='thin', color='AAAAAA')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Helper seguro para formatear fechas sin error de zona horaria
        def format_dt(dt_val, fmt='%d/%m/%Y'):
            if not dt_val:
                return 'N/A'
            try:
                if timezone.is_naive(dt_val):
                    dt_val = timezone.make_aware(dt_val, tz_local)
                dt_local = timezone.localtime(dt_val, tz_local)
                return dt_local.strftime(fmt)
            except Exception:
                return str(dt_val)[:10]

        # ======================================================================
        # HOJA 1: PRÉSTAMOS DEL MES
        # ======================================================================
        ws1 = wb.active
        ws1.title = f"Préstamos {nombre_mes_es} {anio}"

        TOTAL_COLS1 = 16  # A..P

        # Fila 1: Título institucional
        ws1.merge_cells(f'A1:{get_column_letter(TOTAL_COLS1)}1')
        titulo1 = ws1['A1']
        titulo1.value = 'SISTEMA DE GESTIÓN DE PRÉSTAMOS DEPORTIVOS — ULSA'
        titulo1.font    = font_white14
        titulo1.fill    = fill_ulsa
        titulo1.alignment = align_center
        ws1.row_dimensions[1].height = 30

        # Fila 2: Subtítulo con resumen KPI
        ws1.merge_cells(f'A2:{get_column_letter(TOTAL_COLS1)}2')
        subtitulo1 = ws1['A2']
        subtitulo1.value = (
            f'Reporte de Préstamos — {nombre_mes_es} {anio}  |  '
            f'Total: {prestamos.count()}  |  '
            f'Devueltos: {cant_devueltos}  |  '
            f'Activos: {cant_activos}  |  '
            f'Atrasados: {cant_atrasados}  |  '
            f'Pendientes: {cant_pendientes}  |  '
            f'Rechazados: {cant_rechazados}'
        )
        subtitulo1.font    = Font(name='Calibri', bold=True, size=10, color=COLOR_ULSA)
        subtitulo1.alignment = align_center
        ws1.row_dimensions[2].height = 22

        # Fila 3: Separación
        ws1.row_dimensions[3].height = 8

        # Fila 4: Encabezados de columnas
        encabezados1 = [
            'N°', '# Ticket',
            'Fecha Solicitud', 'Hora Solicitud',
            'Fecha Límite Dev.', 'Fecha Real Dev.', 'Hora Devolución',
            'N° Carnet', 'Nombre Solicitante', 'Carrera', 'Año',
            'Equipos Prestados',
            'Entregado por', 'Recibido por',
            'Estado', 'Observaciones'
        ]
        ws1.append([''] * TOTAL_COLS1)  # placeholder fila 3
        for col_idx, enc in enumerate(encabezados1, start=1):
            cell = ws1.cell(row=4, column=col_idx, value=enc)
            cell.fill      = fill_header
            cell.font      = font_white11
            cell.alignment = align_center
            cell.border    = border_all
        ws1.row_dimensions[4].height = 28

        # Filas de datos de préstamos
        fila_inicio1 = 5
        total_tickets  = 0
        total_equipos  = 0

        for idx, p in enumerate(prestamos, start=1):
            es_par = (idx % 2 == 0)
            fill_color = COLOR_FILA_PAR if es_par else COLOR_FILA_IMPAR
            fill_fila = PatternFill(fill_type='solid', start_color=fill_color, end_color=fill_color)

            fecha_sol_fecha = format_dt(p.fecha_prestamo, '%d/%m/%Y')
            fecha_sol_hora  = format_dt(p.fecha_prestamo, '%H:%M')
            fecha_limite    = format_dt(p.fecha_devolucion, '%d/%m/%Y')
            
            if p.estado == 'DEVUELTO':
                fecha_dev_fecha = format_dt(p.fecha_recepcion, '%d/%m/%Y')
                fecha_dev_hora  = format_dt(p.fecha_recepcion, '%H:%M')
            elif p.estado in ('ACTIVO', 'ATRASADO'):
                fecha_dev_fecha = 'En Uso'
                fecha_dev_hora  = '-'
            else:
                fecha_dev_fecha = 'Pendiente'
                fecha_dev_hora  = '-'

            nombre_persona = p.solicitante_externo if p.solicitante_externo else \
                f"{p.estudiante.first_name} {p.estudiante.last_name}".strip() or p.estudiante.username
            carrera_nombre = CARRERAS_MAP.get(p.estudiante.carrera or '', p.estudiante.carrera or 'N/A')
            entregado_por  = p.entregado_por.username if p.entregado_por else 'N/A'
            recibido_por   = p.recibido_por.username if p.recibido_por else 'N/A'

            equipos_lista = ' ; '.join(
                f"{d.equipo.nombre}"
                f"{f' ({d.equipo.marca_modelo})' if d.equipo.marca_modelo else ''}"
                f"{f' [{d.equipo.color}]' if d.equipo.color else ''}"
                f" ×{d.cantidad}"
                for d in p.detalles.all()
            )
            cant_equipos = sum(d.cantidad for d in p.detalles.all())
            total_equipos += cant_equipos

            color_estado = {
                'DEVUELTO':  COLOR_DEVUELTO,
                'PENDIENTE': COLOR_PENDIENTE,
                'ACTIVO':    COLOR_ACTIVO,
                'ATRASADO':  COLOR_ATRASADO,
                'RECHAZADO': COLOR_RECHAZADO,
            }.get(p.estado, 'FFFFFF')

            fila_datos = [
                idx,
                p.id,
                fecha_sol_fecha,
                fecha_sol_hora,
                fecha_limite,
                fecha_dev_fecha,
                fecha_dev_hora,
                p.estudiante.carnet or 'N/A',
                nombre_persona,
                carrera_nombre,
                p.estudiante.ano_cursado or 'N/A',
                equipos_lista,
                entregado_por,
                recibido_por,
                p.estado,
                p.observaciones or '',
            ]

            row_num = fila_inicio1 + idx - 1
            ws1.append(fila_datos)
            ws1.row_dimensions[row_num].height = 22

            for col_idx, val in enumerate(fila_datos, start=1):
                cell = ws1.cell(row=row_num, column=col_idx)
                cell.border    = border_all
                cell.font      = font_normal
                cell.alignment = align_center if col_idx in (1, 2, 3, 4, 5, 6, 7, 10, 11, 14, 15) else align_left

                if col_idx == 15:  # columna Estado con color
                    cell.fill = PatternFill(fill_type='solid', start_color=color_estado, end_color=color_estado)
                    cell.font = Font(name='Calibri', bold=True, size=10)
                else:
                    cell.fill = fill_fila

            total_tickets += 1

        # Fila de totales Hoja 1
        row_totales1 = fila_inicio1 + total_tickets
        ws1.merge_cells(f'A{row_totales1}:P{row_totales1}')
        cell_tot1 = ws1[f'A{row_totales1}']
        cell_tot1.value     = f'TOTAL MES: {total_tickets} préstamos  |  {cant_devueltos} devueltos  |  {cant_activos} activos  |  {cant_atrasados} atrasados  |  {total_equipos} equipos involucrados en total'
        cell_tot1.font      = Font(name='Calibri', bold=True, size=10, color=COLOR_ULSA)
        cell_tot1.fill      = PatternFill(fill_type='solid', start_color='E8F5E9', end_color='E8F5E9')
        cell_tot1.alignment = align_center
        cell_tot1.border    = border_all
        ws1.row_dimensions[row_totales1].height = 22

        anchos1 = [5, 8, 13, 10, 13, 13, 10, 12, 30, 28, 6, 45, 20, 20, 12, 30]
        for i, ancho in enumerate(anchos1, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = ancho

        ws1.freeze_panes = 'A5'

        # ======================================================================
        # HOJA 2: ESTADO DE INVENTARIO EN BODEGA
        # ======================================================================
        ws2 = wb.create_sheet(title="Estado de Inventario")

        TOTAL_COLS2 = 10  # A..J

        # Fila 1: Título institucional
        ws2.merge_cells(f'A1:{get_column_letter(TOTAL_COLS2)}1')
        titulo2 = ws2['A1']
        titulo2.value = 'INVENTARIO GENERAL DE EQUIPOS DEPORTIVOS — ULSA'
        titulo2.font    = font_white14
        titulo2.fill    = fill_ulsa
        titulo2.alignment = align_center
        ws2.row_dimensions[1].height = 30

        # Fila 2: Subtítulo
        ws2.merge_cells(f'A2:{get_column_letter(TOTAL_COLS2)}2')
        subtitulo2 = ws2['A2']
        subtitulo2.value = f'Estado de Bodega de Deportes — Generado: {format_dt(hoy, "%d/%m/%Y %H:%M")}'
        subtitulo2.font    = Font(name='Calibri', bold=True, size=10, color=COLOR_ULSA)
        subtitulo2.alignment = align_center
        ws2.row_dimensions[2].height = 22

        # Fila 3: Separación
        ws2.row_dimensions[3].height = 8

        # Fila 4: Encabezados Hoja 2
        encabezados2 = [
            'N°', 'ID Equipo',
            'Nombre del Equipo', 'Marca / Modelo', 'Color',
            'Cantidad Total', 'Disponible', 'En Préstamo',
            'Estado Stock', 'Descripción'
        ]
        ws2.append([''] * TOTAL_COLS2)  # placeholder fila 3
        for col_idx, enc in enumerate(encabezados2, start=1):
            cell = ws2.cell(row=4, column=col_idx, value=enc)
            cell.fill      = fill_header
            cell.font      = font_white11
            cell.alignment = align_center
            cell.border    = border_all
        ws2.row_dimensions[4].height = 28

        equipos_all = Equipo.objects.all().order_by('nombre')
        sum_total      = 0
        sum_disponibles = 0
        sum_prestados   = 0

        for idx, eq in enumerate(equipos_all, start=1):
            es_par = (idx % 2 == 0)
            fill_color = COLOR_FILA_PAR if es_par else COLOR_FILA_IMPAR
            fill_fila = PatternFill(fill_type='solid', start_color=fill_color, end_color=fill_color)

            en_prestamo = max(0, eq.cantidad_total - eq.cantidad_disponible)
            sum_total       += eq.cantidad_total
            sum_disponibles += eq.cantidad_disponible
            sum_prestados   += en_prestamo

            if eq.cantidad_disponible == 0:
                estado_stock = 'AGOTADO'
                color_stock  = 'FFCDD2'  # rojo
            elif eq.cantidad_disponible <= 2:
                estado_stock = 'STOCK BAJO'
                color_stock  = 'FFF9C4'  # amarillo
            else:
                estado_stock = 'DISPONIBLE'
                color_stock  = 'C8E6C9'  # verde

            fila_inv = [
                idx,
                eq.id,
                eq.nombre,
                eq.marca_modelo or 'N/A',
                eq.color or 'N/A',
                eq.cantidad_total,
                eq.cantidad_disponible,
                en_prestamo,
                estado_stock,
                eq.descripcion or ''
            ]

            row_num = 4 + idx
            ws2.append(fila_inv)
            ws2.row_dimensions[row_num].height = 22

            for col_idx, val in enumerate(fila_inv, start=1):
                cell = ws2.cell(row=row_num, column=col_idx)
                cell.border    = border_all
                cell.font      = font_normal
                cell.alignment = align_center if col_idx in (1, 2, 6, 7, 8, 9) else align_left

                if col_idx == 9:  # columna Estado Stock
                    cell.fill = PatternFill(fill_type='solid', start_color=color_stock, end_color=color_stock)
                    cell.font = Font(name='Calibri', bold=True, size=10)
                else:
                    cell.fill = fill_fila

        # Fila de totales Hoja 2
        row_totales2 = 5 + equipos_all.count()
        ws2.merge_cells(f'A{row_totales2}:J{row_totales2}')
        cell_tot2 = ws2[f'A{row_totales2}']
        cell_tot2.value     = f'TOTAL INVENTARIO: {equipos_all.count()} tipos de equipos  |  {sum_total} unidades registradas en total  |  {sum_disponibles} disponibles  |  {sum_prestados} en préstamo actualmente'
        cell_tot2.font      = Font(name='Calibri', bold=True, size=10, color=COLOR_ULSA)
        cell_tot2.fill      = PatternFill(fill_type='solid', start_color='E8F5E9', end_color='E8F5E9')
        cell_tot2.alignment = align_center
        cell_tot2.border    = border_all
        ws2.row_dimensions[row_totales2].height = 22

        anchos2 = [5, 10, 30, 25, 15, 15, 13, 13, 14, 35]
        for i, ancho in enumerate(anchos2, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = ancho

        ws2.freeze_panes = 'A5'

        # ── Respuesta HTTP usando BytesIO ─────────────────────────────────────
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'Reporte_Prestamos_{nombre_mes_es}_{anio}.xlsx'
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response['X-Filename'] = nombre_archivo
        response['Access-Control-Expose-Headers'] = 'Content-Disposition, X-Filename'
        return response

    except Exception as e:
        logger.exception("Error generando reporte Excel: %s", str(e))
        return Response({'detail': f'Error al generar el reporte Excel: {str(e)}'}, status=500)


# ==========================================
# 3. GOOGLE LOGIN Y PERFIL
# ==========================================

class GoogleLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        token_google = request.data.get('credential')
        if not token_google:
            return Response({'detail': 'Token no proporcionado.'}, status=400)

        try:
            # Validar el token con Google permitiendo un pequeño desfase de reloj (clock skew)
            client_id = getattr(settings, 'GOOGLE_CLIENT_ID', os.getenv('GOOGLE_CLIENT_ID', ''))
            idinfo = id_token.verify_oauth2_token(
                token_google, 
                google_requests.Request(), 
                client_id,
                clock_skew_in_seconds=15
            )

            email = idinfo.get('email', '')
            first_name = idinfo.get('given_name', '')
            last_name = idinfo.get('family_name', '')

            # Verificar dominios permitidos
            dominios_permitidos = ['@est.ulsa.edu.ni', '@ulsa.edu.ni', '@ac.ulsa.edu.ni']
            if not any(email.endswith(dominio) for dominio in dominios_permitidos):
                return Response({'detail': 'Dominio no autorizado. Usa tu correo de la universidad.'}, status=403)

            # Buscar o crear usuario
            user, created = User.objects.get_or_create(username=email, defaults={
                'email': email,
                'first_name': first_name,
                'last_name': last_name
            })

            # Sincronizar flag de admin con la lista blanca
            _sync_admin_flag(user)

            # Generar token DRF
            token, _ = Token.objects.get_or_create(user=user)

            # Verificar si necesita completar perfil (solo estudiantes)
            requiere_perfil = False
            if not user.is_staff and not user.is_superuser:
                if email.endswith('@est.ulsa.edu.ni'):
                    if not user.carnet or not user.carrera or not user.ano_cursado:
                        requiere_perfil = True

            return Response({
                'token': token.key,
                'requiere_completar_perfil': requiere_perfil,
                'user': {
                    'id': str(user.id),
                    'email': user.email,
                    'name': f"{user.first_name} {user.last_name}".strip() or user.username,
                    'role': 'admin' if (user.is_staff or user.is_superuser) else 'student',
                }
            })

        except ValueError as e:
            print(f"Error de Google Auth: {str(e)}")
            return Response({'detail': f'Token inválido: {str(e)}'}, status=401)


import re

class CompletarPerfilView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        carnet = request.data.get('carnet')
        carrera = request.data.get('carrera')
        ano_cursado = request.data.get('ano_cursado')

        if not carnet or not carrera or not ano_cursado:
            return Response({'detail': 'Carnet, carrera y año cursado son obligatorios.'}, status=400)

        carnet_regex = r'^\d{2}-[a-zA-Z0-9\-]{5,}$'
        if not re.match(carnet_regex, carnet):
            return Response({'detail': 'El formato del carnet es inválido.'}, status=400)

        valid_anos = ['1', '2', '3', '4', '5']
        if str(ano_cursado) not in valid_anos:
            return Response({'detail': 'El año cursado debe ser entre 1 y 5.'}, status=400)

        user.carnet = carnet
        user.carrera = carrera
        user.ano_cursado = str(ano_cursado)
        user.save(update_fields=['carnet', 'carrera', 'ano_cursado'])

        return Response({'detail': 'Perfil actualizado correctamente.'})

from django.http import JsonResponse
def clear_broken_images_view(request):
    cleared = []
    for e in Equipo.objects.exclude(imagen=''):
        exists = False
        if e.imagen:
            try:
                exists = os.path.exists(e.imagen.path)
            except Exception:
                pass
        if not exists:
            cleared.append(e.id)
            e.imagen = None
            e.save()
    return JsonResponse({"cleared_ids": cleared})


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def my_ip(request):
    """
    Ruta pública para diagnosticar problemas de red.
    Devuelve la IP pública real que el servidor (Render) está detectando.
    """
    from sgped_api.network_middleware import _get_client_ip
    ip = _get_client_ip(request)
    return Response({
        'ip_detectada': ip,
        'mensaje': 'Usa esta IP para configurar ALLOWED_IPS en Render. También soporta CIDR (ej: 190.212.45.0/24)'
    })