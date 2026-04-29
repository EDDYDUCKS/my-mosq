import os
import openpyxl
from django.core.management.base import BaseCommand
from usuarios.models import Equipo
from django.db import transaction

class Command(BaseCommand):
    help = 'Importa o actualiza el inventario desde el archivo Excel proporcionado'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Ruta absoluta al archivo Excel')

    def handle(self, *args, **kwargs):
        excel_path = kwargs['excel_path']

        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f'El archivo no existe: {excel_path}'))
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        hojas_validas = ['En uso', 'Nuevo']
        
        equipos_procesados = 0
        equipos_creados = 0
        equipos_actualizados = 0

        with transaction.atomic():
            for sheet_name in hojas_validas:
                if sheet_name not in wb.sheetnames:
                    self.stdout.write(self.style.WARNING(f'La hoja "{sheet_name}" no existe en el Excel.'))
                    continue
                
                ws = wb[sheet_name]
                # Encontramos la fila de encabezados (buscando 'Nombre del equipo')
                header_row_idx = None
                for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row[1] and str(row[1]).strip().lower() == 'nombre del equipo':
                        header_row_idx = idx
                        break
                
                if not header_row_idx:
                    self.stdout.write(self.style.ERROR(f'No se encontró la fila de encabezados en la hoja "{sheet_name}".'))
                    continue

                # Procesar datos
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    nombre = row[1]
                    marca_modelo = row[2]
                    color = row[3]
                    cantidad = row[4]

                    # Ignorar filas vacías
                    if not nombre or not cantidad:
                        continue
                    
                    try:
                        cantidad = int(cantidad)
                    except ValueError:
                        continue

                    nombre = str(nombre).strip()
                    # Normalización simple para evitar dobles espacios, etc.
                    nombre = " ".join(nombre.split())

                    marca_modelo = str(marca_modelo).strip() if marca_modelo and str(marca_modelo).strip() not in ['-', 'S/M', 'None'] else None
                    color = str(color).strip() if color and str(color).strip() not in ['-', 'None'] else None

                    # Buscar si el equipo existe (por nombre exacto, marca y color)
                    # Como algunos equipos antiguos no tenían marca o color, podemos buscar primero por nombre
                    # Si ya existe un equipo con ese nombre y sin marca/color definido, podemos actualizar ese.
                    
                    equipo = Equipo.objects.filter(
                        nombre__iexact=nombre,
                        marca_modelo=marca_modelo,
                        color=color
                    ).first()

                    if equipo:
                        # Calcular diferencia para ajustar el disponible
                        diferencia = cantidad - equipo.cantidad_total
                        equipo.cantidad_total = cantidad
                        equipo.cantidad_disponible = max(0, equipo.cantidad_disponible + diferencia)
                        equipo.save()
                        equipos_actualizados += 1
                        self.stdout.write(self.style.SUCCESS(f'Actualizado: {nombre} | Total: {cantidad} | Disp: {equipo.cantidad_disponible}'))
                    else:
                        # Si no existe coincidencia exacta, intentamos ver si existe uno con ese nombre sin atributos (migración de datos viejos)
                        equipo_viejo = Equipo.objects.filter(nombre__iexact=nombre, marca_modelo__isnull=True, color__isnull=True).first()
                        if equipo_viejo:
                            equipo_viejo.marca_modelo = marca_modelo
                            equipo_viejo.color = color
                            diferencia = cantidad - equipo_viejo.cantidad_total
                            equipo_viejo.cantidad_total = cantidad
                            equipo_viejo.cantidad_disponible = max(0, equipo_viejo.cantidad_disponible + diferencia)
                            equipo_viejo.save()
                            equipos_actualizados += 1
                            self.stdout.write(self.style.SUCCESS(f'Migrado y Actualizado: {nombre} | Total: {cantidad}'))
                        else:
                            Equipo.objects.create(
                                nombre=nombre,
                                marca_modelo=marca_modelo,
                                color=color,
                                cantidad_total=cantidad,
                                cantidad_disponible=cantidad
                            )
                            equipos_creados += 1
                            self.stdout.write(self.style.SUCCESS(f'Creado: {nombre} | Total: {cantidad}'))
                    
                    equipos_procesados += 1

        self.stdout.write(self.style.SUCCESS(f'\n--- Resumen ---'))
        self.stdout.write(self.style.SUCCESS(f'Procesados: {equipos_procesados}'))
        self.stdout.write(self.style.SUCCESS(f'Creados: {equipos_creados}'))
        self.stdout.write(self.style.SUCCESS(f'Actualizados: {equipos_actualizados}'))
