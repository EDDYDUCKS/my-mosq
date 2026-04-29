import openpyxl

file_path = r"c:\Users\eddym\OneDrive\Escritorio\Proyectos personales\MOSQ3\Inventario 13032026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n--- Hoja: {sheet_name} ---")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
